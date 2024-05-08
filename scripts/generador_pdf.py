from fpdf import FPDF
from datetime import datetime
from openpyxl.utils.dataframe import dataframe_to_rows
import numpy as np
import pandas as pd 

class PDFWithBackground(FPDF):
    def __init__(self):
        super().__init__()
        self.background = None

    def add_page(self, orientation=''):
        super().add_page(orientation)
        self.create_table()

    def create_table(self):
        self.set_font('helvetica', size= 10)
        
        self.set_y(10)        
        self.cell(40, 20,"", border=1)

        self.image(f'images/hcam1.png',x=11,y=11,w=38,h=18)
        self.set_font('helvetica', 'B', 10)

        self.cell(70, 5,'Formato', "TR",align="C")
        self.set_font('helvetica', size= 10)
        self.cell(70, 5 ,"Código: SGC-ER-FR-007", border="TR")

        self.set_xy(50,15)
        self.cell(70, 5,"", border="R")
        self.cell(70, 5 ,"Versión: 1", border="TR")

        self.set_xy(50,20)
        self.cell(70, 5,"Informe Dosimetría Personal", border="R",align="C")
        self.cell(70, 5 ,"Vigencia: 20/05/2020", border="TR")
        
        x0=self.page_no()
        x1=self.str_alias_nb_pages
        self.set_xy(50,25)
        self.cell(70, 5,"Termoluminiscente", border="RB",align="C")
        self.cell(70, 5 ,f"Página {x0} de {x1}", border="TBR")

    def footer(self):
        yi=-28
        x0=f"Dirección: TIMES SQUARE GARDEN, VIEJO YORK" 
        self.set_y(yi) # Posición a 1.5 cm desde el fondo
        self.set_font('helvetica', 'I', 10) # Configurar la fuente para el pie de página
        self.cell(0, 0, x0, 0, 0, 'L')

        yi+=4
        x1="Edificio Hospital SNOOP DOG "
        self.set_y(yi) # Posición a 1.5 cm desde el fondo
        self.set_font('helvetica', 'I', 10) # Configurar la fuente para el pie de página
        self.cell(0, 0, x1, 0, 0, 'L')

        yi+=4
        x2="Correo: daddy.yankee @reggaeton.com.ec "
        self.set_y(yi) # Posición a 1.5 cm desde el fondo
        self.set_font('helvetica', 'I', 10) # Configurar la fuente para el pie de página
        self.cell(0, 0, x2, 0, 0, 'L')

        yi+=4
        x3="Laboratorio de Dosimetría :1800 88888 Ext. 2296-2295"
        self.set_y(yi) # Posición a 1.5 cm desde el fondo
        self.set_font('helvetica', 'I', 10) # Configurar la fuente para el pie de página
        self.cell(0, 0, x3, 0, 0, 'L')

def header(pdf,n_reporte,tile_hospital,hospital_completo,RUC,f1,f2,fecha):
    
    pdf.set_y(35)
    pdf.set_font('Helvetica',"B",size=12)
    pdf.cell(0,0,f'REPORTE: {n_reporte}_{tile_hospital}',align='C')

    pdf.set_y(40)
    pdf.set_font('Helvetica',size=12)
    pdf.cell(0,0,f'Emisión del informe: {fecha}',align='C')


    pdf.set_y(50)
    pdf.set_font('Helvetica',size=12)
    pdf.cell(0,0,f'DATOS DE LA INSTITUCIÓN USUARIA:',align='L')

    pdf.set_y(55)
    pdf.set_font('Helvetica',"B",size=12)
    pdf.cell(0,0,f'{hospital_completo}',align='L')

    pdf.set_y(60)
    pdf.set_font('Helvetica',size=12)
    pdf.cell(0,0,f'RUC: {RUC}',align='L')

    pdf.set_y(70)
    pdf.set_font('Helvetica',size=12)
    x1=f'Lectura: Bimestral Período: {f1}-{f2}           Tipo: CUERPO ENTERO-TLD'
    pdf.cell(0,0,x1,align='L')

    pdf.set_y(75)
    pdf.set_font('Helvetica',"B",size=8)

    pdf.cell(6, 10 ,"N.-", border="TLRB")
    pdf.cell(74, 10 ,"NOMBRES Y APELLIDOS", border="BTR")
    pdf.cell(18, 10 ,"CEDULA", border="TRB",align="C")
    pdf.cell(20, 5 ,"CODIGO DE", border="TR")
    pdf.cell(20, 5 ,"DOSIS (mSV)", border="TR")
    pdf.cell(25, 5 ,"DOSIS ANUAL", border="TR")
    pdf.cell(25, 5 ,"FIRMA", border="TR")

    pdf.set_xy(108,77.5)
    pdf.cell(20, 7.5 ,"DOSIMETRO", border="BR")
    pdf.cell(20,7.5 ,"Hp(10)", border="BR")
    pdf.cell(25, 5 ,"ACUMULADA", border="R")
    pdf.cell(25,7.5 ,"RECIBIDO", border="BR")

    pdf.set_xy(148,80)
    pdf.cell(25, 5 ,"(mSv) Hp(10)", border="BR")

def tabla_dosis(pdf,df ):
    width_df={0:6,1:74,2:18,
            3:20,4:20,5:25,6:25}
    yi=85
    n_sheet=-1.
    
    pdf.set_font('Helvetica',size=8)

    if df.shape[0]>51:
        n_rows1=50
        n_sheet=(df.shape[0]-n_rows1)//65+1
    
    else: n_rows1=df.shape[0]

    for i in range(n_rows1):
        pdf.set_y(yi)
        for j in range(df.shape[1]):
            if j==1 or j==2: pdf.cell(width_df[j], 3.5 ,f"{df.iloc[i,j]}", border=1)
            else: pdf.cell(width_df[j], 3.5 ,f"{df.iloc[i,j]}", border=1,align="C")
        yi+=3.5    


    if n_sheet>0:
        for k in range(n_sheet):
            nf= n_rows1+65*(k+1) if n_rows1+65*(k+1)<df.shape[0] else df.shape[0]
            ni=n_rows1+65*(k)

            yi=35
            pdf.add_page()
            pdf.set_font('Helvetica',size=8)
            pdf.set_y(yi)
            
            for i in range(ni,nf):
                pdf.set_y(yi)
                for j in range(df.shape[1]):
                    if j==1 or j==2: pdf.cell(width_df[j], 3.5 ,f"{df.iloc[i,j]}", border=1)
                    else: pdf.cell(width_df[j], 3.5 ,f"{df.iloc[i,j]}", border=1,align="C")
                yi+=3.5   
     
    return yi

def acronimos(pdf,yi):
    yi+=5
    ylim=260
    if yi>ylim:
        yi=35
        pdf.add_page()
        
    pdf.set_y(yi)
    pdf.set_font('Helvetica',size=10)
    pdf.cell(92,0,f'      =Dosímetro no Retornado',align='L')
    pdf.cell(45,0,f'      =Dosímetro no Usado',align='L')

    pdf.set_y(yi)
    pdf.set_font('Helvetica',"B")
    pdf.cell(92,0,f'NR',align='L')
    pdf.cell(5,0,f'NU',align='L')

    yi+=5
    pdf.set_y(yi)
    pdf.set_font('Helvetica')
    pdf.cell(80,0,f'      =Dosímetro Dañado no se evalúa la dosis',align='L')
    pdf.cell(80,0,f'                  =Dosímetro Dañado no se evalúa la dosis',align='L')

    pdf.set_y(yi)
    pdf.set_font('Helvetica',"B")
    pdf.cell(90,0,f'DD',align='L')
    pdf.cell(5,0,f'<LD',align='L')
    return yi

def revisado(pdf, fecha,yi):
    x1="Realizado por: ING. JAIR ZEA MARÍN."
    x2="LABORATORIO DE DOSIMETRIA"
    x3="UNIDAD DE SEGURIDAD Y PROTECCION musical"
    x4='HOSPITAL DE ESPECIALIDADES SNOOP DOG'
    
    yi+=30
    ylim=260
    if yi>ylim:
        yi=65
        pdf.add_page()

    pdf.set_font('Helvetica',size=10)
    pdf.set_y(yi)
    pdf.cell(50,0,x1,align='L')

    yi+=5
    pdf.set_y(yi)
    pdf.set_font('Helvetica',"B",size=10)
    pdf.cell(5,0,x2,align='L')

    yi+=5
    pdf.set_y(yi)
    pdf.cell(5,0,x3,align='L')

    yi+=5
    pdf.set_y(yi)
    pdf.cell(5,0,x4,align='L')

    yi+=3
    pdf.set_y(yi)
    pdf.set_font('Helvetica',size=10)
    pdf.cell(25,5,"Revisado por:",align='L', border=1)
    pdf.cell(80,5,"Fis. Méd. FERXXO",align='L', border=1)

    yi+=5
    pdf.set_y(yi)
    pdf.cell(25,5,"Fecha",align='L', border=1)
    pdf.cell(40,5,fecha,align='L', border=1)
    pdf.cell(40,5,"",align='L', border=1)
    return yi

def normas(pdf, yi):
    yi+=8
    ylim=260

    pdf.set_font('Helvetica',"Bu",size=10)
    if yi>ylim:
                yi=35
                pdf.add_page()
    pdf.set_y(yi)
    pdf.cell(5,0,"NOTAS:",align='L')

    pdf.set_font('Helvetica',size=10)

    x2=("La Comisión Internacional de Protección Radiológica, recomienda como límite Hp(10)= 20 mSv/año" 
        "y Hp (0.07) =500mSv/año para el personal POE.")
    x1=x2.split()
    x3=""

    yf=yi+5
    for k in x1:
        if len(x3+" "+k)<=108:
            x3=x3+" "+k 
        
        elif  len(x3+" "+k)>108:
            yi+=5
            if yi>ylim:
                yi=35
                yf=yi
                pdf.add_page()
            
            if yf==yi: 
                pdf.set_xy(12,yi)
                pdf.cell(5,0,"-",align='L')

            pdf.set_xy(15,yi)
            pdf.cell(5,0,x3,align='L')
            x3=" "+k

    yi+=5
    if yi>ylim:
        yi=35
        pdf.add_page()
    pdf.set_xy(15,yi)
    pdf.cell(5,0,x3,align='L')

    yi+=5
    if yi>ylim:
        yi=35
        pdf.add_page()
    
    pdf.set_xy(12,yi)
    pdf.cell(5,0,"-",align='L')
    x2="Las dosis superiores a 1.5mSv/mes a cuerpo entero deben ser reportadas  al MEM-SCAN."
    pdf.set_xy(15,yi)
    pdf.cell(5,0,x2,align='L')

    yi+=5
    if yi>ylim:
        yi=35
        pdf.add_page()
    
    
    x2=("Según Regulación Internacional los valores de dosis menores a 0,1mSv son considerados" 
       "CERO para efectos de dosis acumuladas")
    x1=x2.split()
    x3=""

    yf=yi+5
    for k in x1:
        if len(x3+" "+k)<=108:
            x3=x3+" "+k 
        
        elif  len(x3+" "+k)>108:
            yi+=5
            if yi>ylim:
                yi=35
                yf=yi
                pdf.add_page()
            
            if yf==yi: 
                pdf.set_xy(12,yi)
                pdf.cell(5,0,"-",align='L')

            pdf.set_xy(15,yi)
            pdf.cell(5,0,x3,align='L')
            x3=" "+k

    yi+=5
    if yi>ylim:
        yi=35
        pdf.add_page()
    pdf.set_xy(15,yi)
    pdf.cell(5,0,x3,align='L')
        


    x2=("El dosímetro es personal y debe manejar con responsabilidad lo cual debe ser usado solo" 
        "en horas de trabajo, además debe ser utilizado únicamente dentro de la Institución."    )
    x1=x2.split()
    x3=""

    yf=yi+5
    for k in x1:
        if len(x3+" "+k)<=108:
            x3=x3+" "+k 
        
        elif  len(x3+" "+k)>108:
            yi+=5
            if yi>ylim:
                yi=35
                yf=yi
                pdf.add_page()
            
            if yf==yi: 
                pdf.set_xy(12,yi)
                pdf.cell(5,0,"-",align='L')

            pdf.set_xy(15,yi)
            pdf.cell(5,0,x3,align='L')
            x3=" "+k

    yi+=5
    if yi>ylim:
        yi=35
        pdf.add_page()
    pdf.set_xy(15,yi)
    pdf.cell(5,0,x3,align='L')


    
    x2=("Las Normas Básicas de Seguridad y Protección  Radiológica, exponen los requisitos" 
        "fundamentales que debe  cumplir toda actividad que implique exposición a las radiaciones" 
        "ionizantes, siendo estos los más importantes para la exposición externa: tiempo, blindaje y"
        "distancia, estos tres parámetros son fundamentales al momento de la práctica donde se genere o" 
        "se emita radiación ionizante.")
    x1=x2.split()
    x3=""

    yf=yi+5
    for k in x1:
        if len(x3+" "+k)<=108:
            x3=x3+" "+k 
        
        elif  len(x3+" "+k)>108:
            yi+=5
            if yi>ylim:
                yi=35
                yf=yi
                pdf.add_page()
            
            if yf==yi: 
                pdf.set_xy(12,yi)
                pdf.cell(5,0,"-",align='L')

            pdf.set_xy(15,yi)
            pdf.cell(5,0,x3,align='L')
            x3=" "+k

    yi+=5
    if yi>ylim:
        yi=35
        pdf.add_page()
    pdf.set_xy(15,yi)
    pdf.cell(5,0,x3,align='L')

    x2=("NORMATIVA 1 PARA EL DESARROLLO E INSCRIPCION DE LOS ARGUMENTOS")
    x1=x2.split()
    x3=""

    yf=yi+5
    for k in x1:
        if len(x3+" "+k)<=108:
            x3=x3+" "+k 
        
        elif  len(x3+" "+k)>108:
            yi+=5
            if yi>ylim:
                yi=35
                yf=yi
                pdf.add_page()
            
            if yf==yi: 
                pdf.set_xy(12,yi)
                pdf.cell(5,0,"-",align='L')

            pdf.set_xy(15,yi)
            pdf.cell(5,0,x3,align='L')
            x3=" "+k

    yi+=5
    if yi>ylim:
        yi=35
        pdf.add_page()
    pdf.set_xy(15,yi)
    pdf.cell(5,0,x3,align='L')
    


    return yi

def elimina_strin(x):
    try:
        x=float(x)
    
    except:
        x=np.NaN
    return x



def df_to_pdf(df,fecha_canje, n_reporte,tile_hospital,hospital_completo,RUC,f1,f2):
    name=df.columns.to_list()

    total=[t for t in name if "TOTAL" in t ]
    total=total[0]
    hp="DOSIS (mSv) Hp(10)"

    df["FIRMA RECIBIDO"]=""
    df[hp]=df[fecha_canje].round(3)
    df[total]=df[total].round(3)

    df[hp]=df[hp].apply(elimina_strin)
    df[total]=df[total].apply(elimina_strin)

    df=df.dropna(subset=["N.-","NOMBRES Y APELLIDOS", "CEDULA","CODIGO DE DOSIMETRO"])


    df=df[["N.-","NOMBRES Y APELLIDOS", "CEDULA","CODIGO DE DOSIMETRO", 
        hp,total,"FIRMA RECIBIDO"]]

    df = df.replace(np.nan, "")

    return create_pdf(df, n_reporte,tile_hospital,hospital_completo,RUC,f1,f2)


def create_pdf(df, n_reporte,tile_hospital,hospital_completo,RUC,f1,f2):
    fecha_today = datetime.today().strftime("%d/%m/%Y")

    pdf = PDFWithBackground()

    pdf.add_page()

    header(pdf,n_reporte,tile_hospital,hospital_completo,RUC,f1,f2,fecha_today)

    yi=tabla_dosis(pdf,df)

    yi=acronimos(pdf,yi)

    yi=normas(pdf,yi)

    yi=revisado(pdf, fecha_today,yi)
    
    pdf.output('Mi_primer_pdf.pdf')

    return pdf 
    