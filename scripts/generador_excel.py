from openpyxl import Workbook
from openpyxl.styles import Font, Color, Alignment, PatternFill, Border
from openpyxl.drawing.image import Image
import string
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook import Workbook
import pandas as pd
import numpy as np
from io import BytesIO

def descargar_excel(wb):
    excel_bytes = BytesIO()
    wb.save(excel_bytes)
    excel_bytes.seek(0)
    return excel_bytes

def excel_asistencia(df,title):
    name=df.columns.to_list()
    # Crear un nuevo libro y seleccionar la hoja activa
    wb = Workbook()
    ws = wb.active
    # Añadir texto a una celda
    ws['A1'] = title

    # Establecer fuente, tamaño y color
    ws['A1'].font = Font(name='Arial',size=13,bold=True, underline="single" )
    ws.merge_cells('A1:D2')  # Combinar celdas desde E1 hasta G2

    # Aplicar formato de alineación al texto combinado
    ws['A1'].alignment = Alignment(horizontal='center', vertical='bottom')

    ws.row_dimensions[4].height = 15
    lecture=25
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 13
    for k in  range(4,len(name)):
        letra=string.ascii_uppercase[k]
        ws.column_dimensions[letra].width = lecture


    ## REALIZO UNA COPIA DEL ENCABEZADO DEL DOSIMETRO EN EXCEL
    for k in  range(len(name)):
        letra=string.ascii_uppercase[k]+"4"
        ws[letra]=name[k]
        ws[letra].font = Font(name='Arial',size=10,bold=True, italic=True,underline="single" )
        ws[letra].alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)

    # LLENO LAS CELDAS CON LA NUEVA INFORMACION 
    for r in dataframe_to_rows(df, index=False, header=False):
        ws.append(r)
        
        for cell in ws[ws.max_row]:
            cell.font = Font(name='Arial',size=10)  # Cambiar tamaño de fuente

            if "nuevo" in str(cell.value) :
                for k in  range(len(name)):
                    letra=string.ascii_uppercase[k]+str(cell.row)
                    ws[letra].fill = PatternFill(fgColor= "8cff66",fill_type='solid')       

            if ".C" in str(cell): cell.alignment = Alignment(horizontal='left', vertical='center')     
            else: cell.alignment = Alignment(horizontal='center', vertical='center') 



    ws.sheet_view.zoomScale=80
    #ws.sheet_view = SheetView(zoomScale=80) ##Guardo el excel con una ventana de zoom al 80%
    return wb

def excel_historial_dosimetros(df):
    name=df.columns.to_list()
    # Crear un nuevo libro y seleccionar la hoja activa
    wb = Workbook()
    ws = wb.active
    # Añadir texto a una celda
    ws['A1'] = f"HISTORIAL DE DOSIMETROS POE UNIDADES INTERNAS 2024"

    # Establecer fuente, tamaño y color
    ws['A1'].font = Font(name='Arial',size=13,bold=True, underline="single" )
    ws.merge_cells('A1:G2')  # Combinar celdas desde E1 hasta G2

    # Aplicar formato de alineación al texto combinado
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    ws.row_dimensions[4].height = 65
    lecture=13
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 13

    for k in range(4, len(name)-1):
        letra=string.ascii_uppercase[k]
        ws.column_dimensions[letra].width = lecture
    letra=string.ascii_uppercase[k+1]
    ws.column_dimensions[letra].width = 30
    
## REALIZO UNA COPIA DEL ENCABEZADO DEL DOSIMETRO EN EXCEL
    for k in  range(len(name)):
        letra=string.ascii_uppercase[k]+"4"
        ws[letra]=name[k]
        ws[letra].font = Font(name='Arial',size=10,bold=True, italic=True,underline="single" )
        ws[letra].alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)

    # LLENO LAS CELDAS CON LA NUEVA INFORMACION 
    for r in dataframe_to_rows(df, index=False, header=False):
        ws.append(r)
        
        for cell in ws[ws.max_row]:
            cell.font = Font(name='Arial',size=10)  # Cambiar tamaño de fuente

            if "nuevo" in str(cell.value) :
                for k in  range(len(name)):
                    letra=string.ascii_uppercase[k]+str(cell.row)
                    ws[letra].fill = PatternFill(fgColor= "8cff66",fill_type='solid')       

            if ".C" in str(cell): cell.alignment = Alignment(horizontal='left', vertical='center')     
            else: cell.alignment = Alignment(horizontal='center', vertical='center') 



    ws.sheet_view.zoomScale=80
    return wb


def excel_reporte_lectura_dosis(df):
    name=df.columns.to_list()
    # Crear un nuevo libro y seleccionar la hoja activa
    wb = Workbook()
    ws = wb.active
    # Añadir texto a una celda
    ws['A1'] = f"DOSIS DOSIMETRICA ACUMULADA POE UNIDADES EXTERNAS 2024"

    # Establecer fuente, tamaño y color
    ws['A1'].font = Font(name='Arial',size=13,bold=True, underline="single" )
    ws.merge_cells('A1:L2')  # Combinar celdas desde E1 hasta G2

    # Aplicar formato de alineación al texto combinado
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    ws.row_dimensions[4].height = 65
    lecture=13
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 13
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = lecture
    ws.column_dimensions['G'].width = lecture
    ws.column_dimensions['H'].width = lecture
    ws.column_dimensions['I'].width = lecture
    ws.column_dimensions['J'].width = lecture
    ws.column_dimensions['K'].width = lecture
    ws.column_dimensions['L'].width = 14
    ws.column_dimensions['M'].width = 40

    ## REALIZO UNA COPIA DEL ENCABEZADO DEL DOSIMETRO EN EXCEL
    for k in  range(13):
        letra=string.ascii_uppercase[k]+"4"
        ws[letra]=name[k]
        if name[k].startswith("LECTURA") or name[k].startswith("TOTAL") :
            ws[letra].font = Font(name='Arial',size=10,bold=True)
            ws[letra].alignment = Alignment(horizontal='left', vertical='center',wrap_text=True)
        else: 
            ws[letra].font = Font(name='Arial',size=10,bold=True, italic=True,underline="single" )
            ws[letra].alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)

    # LLENO LAS CELDAS CON LA NUEVA INFORMACION 
    for r in dataframe_to_rows(df, index=False, header=False):
        ws.append(r)
        for cell in ws[ws.max_row]:
            if ".C" in str(cell): cell.alignment = Alignment(horizontal='left', vertical='center')     
            else: cell.alignment = Alignment(horizontal='center', vertical='center') 
            cell.font = Font(name='Arial',size=10)  # Cambiar tamaño de fuente

    # Aplicar un color de fondo a una celda
    #ws['A1'].fill = PatternFill(start_color='050505',end_color='050505',fill_type='solid')

    ws.sheet_view.zoomScale=80
    return wb